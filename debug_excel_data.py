#!/usr/bin/env python3
"""
Debug what's actually in the Excel file to fix the agent
"""

import os
import sys

def debug_excel_without_pandas():
    """Debug Excel file without pandas dependency"""
    print("🔍 DEBUGGING EXCEL FILE STRUCTURE")
    print("=" * 60)
    
    excel_path = "data/excel/Finaptive PBI Mining Data Set.xlsx"
    
    if not os.path.exists(excel_path):
        print(f"❌ Excel file not found: {excel_path}")
        return False
    
    print(f"✅ Excel file found: {excel_path}")
    print(f"📏 File size: {os.path.getsize(excel_path):,} bytes")
    print("")
    
    # Try to use a different approach to read Excel
    try:
        import openpyxl
        print("📊 Using openpyxl to examine Excel structure...")
        
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        print(f"📋 Worksheets found: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n📄 Worksheet: {sheet_name}")
            print(f"   📐 Dimensions: {ws.max_row} rows x {ws.max_column} columns")
            
            # Get column headers (first row)
            headers = []
            for col in range(1, min(ws.max_column + 1, 11)):  # First 10 columns
                cell_value = ws.cell(row=1, column=col).value
                headers.append(str(cell_value) if cell_value else f"Col{col}")
            
            print(f"   📝 Headers: {headers}")
            
            # Sample some data rows
            print(f"   📊 Sample data:")
            for row in range(2, min(6, ws.max_row + 1)):  # Rows 2-5
                row_data = []
                for col in range(1, min(6, ws.max_column + 1)):  # First 5 columns
                    cell_value = ws.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value else "")
                print(f"      Row {row}: {row_data}")
            
            # Look for Ontario/region data
            ontario_found = False
            year_2023_found = False
            
            print(f"   🔍 Searching for 'Ontario' and '2023'...")
            for row in range(1, min(101, ws.max_row + 1)):  # Check first 100 rows
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        cell_str = str(cell_value).lower()
                        if 'ontario' in cell_str:
                            ontario_found = True
                            print(f"      ✅ Found 'Ontario' at Row {row}, Col {col}: {cell_value}")
                        if '2023' in cell_str:
                            year_2023_found = True
                            print(f"      ✅ Found '2023' at Row {row}, Col {col}: {cell_value}")
            
            if not ontario_found:
                print(f"      ❌ No 'Ontario' found in first 100 rows")
            if not year_2023_found:
                print(f"      ❌ No '2023' found in first 100 rows")
        
        wb.close()
        return True
        
    except ImportError:
        print("❌ openpyxl not available, trying alternative approach...")
        return debug_excel_alternative()
    except Exception as e:
        print(f"❌ Error reading Excel: {e}")
        return debug_excel_alternative()

def debug_excel_alternative():
    """Alternative debugging without special libraries"""
    print("\n🔍 ALTERNATIVE EXCEL DEBUGGING")
    print("=" * 60)
    
    excel_path = "data/excel/Finaptive PBI Mining Data Set.xlsx"
    
    # Try to see if we can detect file format issues
    try:
        with open(excel_path, 'rb') as f:
            # Read first few bytes to check file signature
            header = f.read(8)
            print(f"📄 File header: {header.hex()}")
            
            # Excel files should start with specific signatures
            if header[:2] == b'PK':
                print("✅ File appears to be a ZIP-based format (modern Excel)")
            elif header[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
                print("✅ File appears to be legacy Excel format")
            else:
                print("⚠️ File format may be unusual")
            
            f.seek(0, 2)  # Seek to end
            file_size = f.tell()
            print(f"📏 Total file size: {file_size:,} bytes")
            
    except Exception as e:
        print(f"❌ Could not read file: {e}")
        return False
    
    return True

def suggest_fixes():
    """Suggest fixes for the agent"""
    print("\n🔧 SUGGESTED FIXES FOR THE AGENT")
    print("=" * 60)
    
    print("**Problem**: Agent says 'Ontario 2023 data not found'")
    print("**Likely Causes**:")
    print("   1. 📋 Agent looking at wrong worksheet")
    print("   2. 🔤 Data format doesn't match expectations")
    print("   3. 📅 Date stored as number, not text '2023'")
    print("   4. 🌍 Region stored differently ('ON' vs 'Ontario')")
    print("   5. 📊 Data in different columns than expected")
    print("")
    
    print("**Fixes to implement**:")
    print("   ✅ Make agent examine ALL worksheets")
    print("   ✅ Add fuzzy matching for regions")
    print("   ✅ Handle different date formats")
    print("   ✅ Show agent what data it actually found")
    print("   ✅ Add data discovery before analysis")
    print("")
    
    print("**Agent improvements needed**:")
    print("   1. 🔍 Data discovery phase")
    print("   2. 📋 Show sample data to user")
    print("   3. 🤖 Smarter data matching")
    print("   4. 🔄 Multiple search strategies")
    print("   5. 📊 Better error reporting")

def create_fix_plan():
    """Create a plan to fix the agent"""
    print("\n📋 AGENT FIX PLAN")
    print("=" * 60)
    
    fixes = [
        "Add data discovery phase to show what's actually in the Excel",
        "Implement fuzzy matching for regions (Ontario, ON, ontario)",
        "Handle different date formats (2023, '2023', date objects)",
        "Make agent examine all worksheets, not just first one",
        "Add verbose logging so we can see what agent is doing",
        "Implement fallback search strategies",
        "Show sample data to user when data not found",
        "Add data validation before analysis"
    ]
    
    for i, fix in enumerate(fixes, 1):
        print(f"   {i}. {fix}")
    
    print(f"\n🎯 Priority: Fix data discovery first!")

if __name__ == "__main__":
    success = debug_excel_without_pandas()
    suggest_fixes()
    create_fix_plan()
    
    print(f"\n" + "=" * 60)
    print("🔧 NEXT STEPS")
    print("=" * 60)
    if success:
        print("✅ Excel file is readable")
        print("🔧 Need to fix agent data discovery logic")
        print("📊 Add better worksheet examination")
        print("🔍 Implement smarter data matching")
    else:
        print("❌ Excel file has issues")
        print("🔧 Fix file access first, then agent logic")